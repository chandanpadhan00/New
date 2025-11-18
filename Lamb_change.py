import boto3, psycopg2, json, io, datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment


# ==============================
# NEW HELPER: Newly Flagged Cases
# ==============================
def get_newly_flagged_cases(cursor, vendor_id):
    """
    Returns rows + headers for cases that were flagged in the latest run
    but not in the previous run, for the given vendor_id.
    """
    query = """
        WITH process_times AS (
            SELECT DISTINCT process_datetime
            FROM asnfdm.wkly_flagged_cases_hist
            WHERE vendor_id = %s
            ORDER BY process_datetime DESC
            LIMIT 2
        ),
        max_times AS (
            SELECT
                MAX(process_datetime) AS max_time,
                MIN(process_datetime) AS second_max_time
            FROM process_times
        )
        SELECT
            a.vendor_id,
            a.patient_id,
            a.case_id,
            a.qc_rule_no
        FROM asnfdm.wkly_flagged_cases_hist a, max_times mt
        WHERE a.process_datetime = mt.max_time
          AND a.vendor_id = %s
          AND NOT EXISTS (
              SELECT 1
              FROM asnfdm.wkly_flagged_cases_hist b, max_times mt2
              WHERE b.process_datetime = mt2.second_max_time
                AND b.case_id = a.case_id
                AND b.vendor_id = a.vendor_id
          );
    """
    cursor.execute(query, (vendor_id, vendor_id))
    rows = cursor.fetchall()
    headers = [col[0] for col in cursor.description]
    return rows, headers


# =================
# REPORT GENERATION
# =================
def load_report_to_s3(connection_string, source_name, bucket, csv_file, env, batch_id):
    """
    Builds the QC / exception Excel report and uploads it to S3.

    Sheets:
      1. sheet_1  -> detailed failed records (PCD 2.0 Exception)
      2. sheet_2  -> rule-wise counts (COUNT)
      3. sheet_3  -> newly flagged cases vs previous run
    """
    s3_client = boto3.client('s3')
    conn = psycopg2.connect(connection_string)
    cursor = conn.cursor()

    try:
        # ---------- Read config ----------
        config = json.load(open("conf.json", "r"))
        qc_conf = config['conf_json']['qc_details']

        report_bucket = qc_conf['qc_bucket']

        # Sheet names from config
        sheet_1_cfg = qc_conf['sheet_1']          # can be dict or string
        sheet_2_cfg = qc_conf['sheet_2']
        sheet_3_cfg = qc_conf.get('sheet_3', 'Newly Flagged Cases')

        sheet_name_1 = sheet_1_cfg[source_name] if isinstance(sheet_1_cfg, dict) else sheet_1_cfg
        sheet_name_2 = sheet_2_cfg[source_name] if isinstance(sheet_2_cfg, dict) else sheet_2_cfg
        sheet_name_3 = sheet_3_cfg

        # ---------- Vendor–specific details ----------
        if source_name == 'SNX':
            report_table   = qc_conf['snx_qc_table']
            report_columns = qc_conf['snx_qc_columns']
            vendor_qc_conf = qc_conf['snx_qc_details'][env]
        elif source_name == 'REGALORX':
            report_table   = qc_conf['regalorx_qc_table']
            report_columns = qc_conf['regalorx_qc_columns']
            vendor_qc_conf = qc_conf['regalorx_qc_details'][env]
        else:
            raise Exception(f"Unsupported vendor/source_name: {source_name}")

        qc_outbound_path = vendor_qc_conf['qc_outbound_path']
        qc_report_path   = vendor_qc_conf['csv_file_path'].replace('{batch_id}', str(batch_id))

        print(f"QC report S3 key: s3://{report_bucket}/{qc_report_path}")
        print(f"QC outbound S3 key: s3://{report_bucket}/{qc_outbound_path}")

        # ---------- SHEET 1: detailed failed rows ----------
        # You can tweak WHERE / ORDER BY if your existing logic is slightly different
        query1 = f"""
            SELECT DISTINCT {report_columns}
            FROM {report_table}
            WHERE qc_rule_no <> 0
            ORDER BY qc_rule_no;
        """
        cursor.execute(query1)
        rows1 = cursor.fetchall()
        headers1 = [col[0] for col in cursor.description]

        # ---------- SHEET 2: rule-wise counts ----------
        query2 = f"""
            SELECT
                qc_rule_no,
                qc_rule,
                COUNT(*) AS record_count
            FROM {report_table}
            WHERE qc_rule_no <> 0
            GROUP BY qc_rule_no, qc_rule
            ORDER BY qc_rule_no;
        """
        cursor.execute(query2)
        rows2 = cursor.fetchall()
        headers2 = [col[0] for col in cursor.description]

        # (Optional) Insert into history table using rows2
        current_week = datetime.datetime.now().strftime("%W %Y")
        for row in rows2:
            qc_rule_no, qc_rule, counts = row
            insert_query = (
                "INSERT INTO anfsdm.pap_qc_rule_history "
                "(batch_id, source_name, qc_rule_no, qc_rule, record_count, week_label) "
                "VALUES ({}, '{}', {}, $$${}$$, {}, '{}')"
            ).format(batch_id, source_name, qc_rule_no, qc_rule, counts, current_week)
            # If you already have execute_query() helper, you can call it here.
            # execute_query(insert_query)

        # ---------- SHEET 3: newly flagged cases ----------
        new_rows, new_headers = get_newly_flagged_cases(cursor, source_name)

        # ---------- Build DataFrames ----------
        df1 = pd.DataFrame(rows1, columns=headers1)
        df2 = pd.DataFrame(rows2, columns=headers2)
        df3 = pd.DataFrame(new_rows, columns=new_headers)

        # ---------- Write to Excel (in memory) ----------
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine="xlsxwriter") as writer:
            df1.to_excel(writer, sheet_name=sheet_name_1, index=False)
            df2.to_excel(writer, sheet_name=sheet_name_2, index=False)
            df3.to_excel(writer, sheet_name=sheet_name_3, index=False)

        buffer.seek(0)

        # ---------- Style headers with openpyxl ----------
        workbook = load_workbook(buffer)

        sheet1 = workbook[sheet_name_1]
        sheet2 = workbook[sheet_name_2]
        sheet3 = workbook[sheet_name_3]

        for ws in (sheet1, sheet2, sheet3):
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="left")

        modified_buffer = io.BytesIO()
        workbook.save(modified_buffer)
        modified_buffer.seek(0)

        # ---------- Upload to S3 & copy to outbound ----------
        s3_client.upload_fileobj(modified_buffer, report_bucket, qc_report_path)
        print(f"Uploaded report to S3: s3://{report_bucket}/{qc_report_path}")

        s3_client.copy_object(
            Bucket=report_bucket,
            Key=qc_outbound_path,
            CopySource={"Bucket": report_bucket, "Key": qc_report_path},
        )
        print(f"Copied report to outbound: s3://{report_bucket}/{qc_outbound_path}")

        conn.commit()
        return qc_outbound_path, bucket

    except Exception as e:
        print(f"Error while generating report: {e}")
        conn.rollback()
        raise
    finally:
        cursor.close()
        conn.close()
